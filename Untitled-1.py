import pandas as pd
from openpyxl import Workbook

# %%

#price라는 Series 생성
price = pd.Series([500, 1000, 2000, 4000])
print("Series type:")
print(type(price)) #Series는 pandas의 클래스


# %%
print("Series 전체 출력")
print(price)








# %%
print("Series 인덱싱")
print(price[0]) 
print(price[1])

price.to_excel('test.xlsx')

# 엑셀파일 쓰기
write_wb=Workbook()

#이름이 있는 시트를 생성
write_ws=write_wb.create_sheet('생성시트')

#Sheet1에다 입력
write_ws=write_wb.active
write_ws['A1']='숫자'

#행 단위로 추가
write_ws.append([1,2,3])

#셀 단위로 추가
write_ws.cell(5,5,'5행5열')
write_wb.save("test2.xlsx")

# %%